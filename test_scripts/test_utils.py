from statistics import mode

import pandas as pd
import os
from datetime import datetime
import openpyxl

def load_test_cases_csv(csv_path):
   data =  pd.read_csv(csv_path,on_bad_lines='skip').to_dict('records')
   print(data)
   return data
def load_test_cases(excel_path):
    return pd.read_excel(excel_path).to_dict('records')
def prepare_summary(no_of_iters):
    return {
        'Question': [],
        'run_date': [],
        f'total_time_for_{no_of_iters}_iterations': [],
        f'PASS_count_out_of_{no_of_iters}': [],
        f'FAIL_count_out_of_{no_of_iters}': [],
        'overall_status': []
    }

def update_summary(summary, test_case_id, no_of_iters, pass_count, fail_count, start_time, end_time):
    if test_case_id in summary['Question']:
        index = summary['Question'].index(test_case_id)
        summary['overall_status'][index] = 'PASS' if fail_count == 0 else 'FAIL'
        summary[f'PASS_count_out_of_{no_of_iters}'][index] = pass_count
        summary[f'FAIL_count_out_of_{no_of_iters}'][index] = fail_count
        summary[f'total_time_for_{no_of_iters}_iterations'][index] = str(end_time - start_time)
        summary['run_date'][index] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
    else:
        summary['Question'].append(test_case_id)
        summary['overall_status'].append('PASS' if fail_count == 0 else 'FAIL')
        summary[f'PASS_count_out_of_{no_of_iters}'].append(pass_count)
        summary[f'FAIL_count_out_of_{no_of_iters}'].append(fail_count)
        summary[f'total_time_for_{no_of_iters}_iterations'].append(str(end_time - start_time))
        summary['run_date'].append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    return summary

def save_test_result(df, file_path, test_case_id):
    if os.path.exists(file_path):
        mode = "a"
        with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=f"{test_case_id}_test_result", index=False)

    else:
        mode = "w"
        with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode) as writer:
            df.to_excel(writer, sheet_name=f"{test_case_id}_test_result", index=False)

def save_summary(summary, file_path):
    df1 = pd.DataFrame(summary)
    
    if os.path.exists(file_path): 
        mode = "a" 
        # lets position the summary sheet at the beginning of the excel file

        with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, if_sheet_exists='replace') as writer:
            df1.to_excel(writer, sheet_name="summary", index=False)
        # Reorder sheets: move 'summary' to the beginning
        wb = openpyxl.load_workbook(file_path)
        if 'summary' in wb.sheetnames:
            ws = wb['summary']
            wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index('summary')))
            wb.save(file_path)
    else:
        mode = "w"
        with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode) as writer:
            df1.to_excel(writer, sheet_name="summary", index=False)